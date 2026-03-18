# app.py
import io
from datetime import datetime, timedelta

import duckdb
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import DATE_FORMATS, PARQUET_MESTRE

# ── Página ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Detector de Duplicidades de OS",
    page_icon="🔍",
    layout="wide",
)

# ── DuckDB helpers ────────────────────────────────────────────────────────────
def get_con():
    return duckdb.connect()


@st.cache_data(show_spinner="⚡ Lendo base histórica via DuckDB...")
def carregar_base(
    data_ini: str | None = None,
    data_fim: str | None = None,
    lotes: list[str] | None = None,
) -> pd.DataFrame:
    if not PARQUET_MESTRE.exists():
        return pd.DataFrame()

    con   = get_con()
    where = []
    if data_ini:
        where.append(f"_data_parsed >= DATE '{data_ini}'")
    if data_fim:
        where.append(f"_data_parsed <= DATE '{data_fim}'")
    if lotes:
        lotes_str = ", ".join(f"'{l}'" for l in lotes)
        where.append(f"_lote IN ({lotes_str})")

    where_clause = f"WHERE {' AND '.join(where)}" if where else ""
    query = f"""
        SELECT *
        FROM read_parquet('{PARQUET_MESTRE}')
        {where_clause}
        ORDER BY _data_parsed
    """
    df = con.execute(query).df()
    con.close()
    return df


@st.cache_data(show_spinner="📋 Carregando lotes disponíveis...")
def listar_lotes() -> list[str]:
    if not PARQUET_MESTRE.exists():
        return []
    con   = get_con()
    query = f"SELECT DISTINCT _lote FROM read_parquet('{PARQUET_MESTRE}') ORDER BY _lote"
    lotes = con.execute(query).df()["_lote"].tolist()
    con.close()
    return lotes


@st.cache_data(show_spinner="📊 Calculando estatísticas da base...")
def stats_base() -> dict:
    if not PARQUET_MESTRE.exists():
        return {}
    con   = get_con()
    query = f"""
        SELECT
            COUNT(*)                       AS total_registros,
            COUNT(DISTINCT _lote)          AS total_lotes,
            MIN(_data_parsed)::VARCHAR     AS data_min,
            MAX(_data_parsed)::VARCHAR     AS data_max,
            COUNT(DISTINCT _cliente_norm)  AS clientes_unicos,
            COUNT(DISTINCT _servico_norm)  AS servicos_unicos
        FROM read_parquet('{PARQUET_MESTRE}')
    """
    row = con.execute(query).df().iloc[0].to_dict()
    con.close()
    return row


# ── Parse de datas robusto ────────────────────────────────────────────────────
def parse_dates_robust(series: pd.Series) -> pd.Series:
    result    = pd.Series([pd.NaT] * len(series), index=series.index)
    remaining = series.copy()
    for fmt in DATE_FORMATS:
        mask = result.isna() & remaining.notna()
        if not mask.any():
            break
        parsed = pd.to_datetime(remaining[mask], format=fmt, errors="coerce")
        result[mask] = parsed
    still_null = result.isna() & series.notna()
    if still_null.any():
        result[still_null] = pd.to_datetime(
            series[still_null], infer_datetime_format=True, errors="coerce"
        )
    return result


# ── Detecção de duplicidades ──────────────────────────────────────────────────
@st.cache_data(show_spinner="🔎 Detectando duplicidades...")
def detect_duplicates(
    df: pd.DataFrame,
    col_cliente: str,
    col_servico: str,
    col_data: str,
    janela_dias: int,
    col_os: str | None = None,
    cols_extras: tuple[str, ...] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:

    work = df.copy()

    # ── Datas ────────────────────────────────────────────────────────────────
    if "_data_parsed" in work.columns:
        work["_data_parsed"] = pd.to_datetime(work["_data_parsed"], errors="coerce")
    else:
        work["_data_parsed"] = parse_dates_robust(work[col_data].astype(str).str.strip())

    total     = len(work)
    invalidas = work["_data_parsed"].isna().sum()
    validas   = total - invalidas

    with st.expander("📅 Diagnóstico de datas", expanded=invalidas > 0):
        c1, c2, c3 = st.columns(3)
        c1.metric("Total de registros",       f"{total:,}")
        c2.metric("Datas reconhecidas",        f"{validas:,}")
        c3.metric("Datas inválidas/ignoradas", f"{invalidas:,}",
                  delta=f"-{invalidas}" if invalidas else None,
                  delta_color="inverse")
        if invalidas > 0:
            exemplos = work[work["_data_parsed"].isna()][col_data].dropna().unique()[:10]
            st.warning(f"Exemplos não reconhecidos: `{'`, `'.join(map(str, exemplos))}`")
        else:
            amostra = work[[col_data, "_data_parsed"]].dropna().head(5).copy()
            amostra["_data_parsed"] = amostra["_data_parsed"].dt.strftime("%d/%m/%Y")
            amostra.columns = ["Valor original", "Interpretado como"]
            st.success("Todas as datas foram reconhecidas.")
            st.dataframe(amostra, use_container_width=True)

    # ── Remove linhas sem data, sem cliente ou sem serviço ───────────────────
    work = work.dropna(subset=["_data_parsed"]).copy()
    work = work[
        work[col_cliente].notna() & (work[col_cliente].astype(str).str.strip() != "") &
        work[col_servico].notna() & (work[col_servico].astype(str).str.strip() != "")
    ].copy()

    if work.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    work = work.reset_index(drop=True)
    work["_row_id"]       = work.index
    work["_cliente_norm"] = work[col_cliente].astype(str).str.strip().str.upper()
    work["_servico_norm"] = work[col_servico].astype(str).str.strip().str.upper()

    # ── Algoritmo de detecção ─────────────────────────────────────────────────
    # ATENÇÃO: grupo_counter é GLOBAL — não reinicia a cada (cliente, serviço)
    # Isso garante que grupo_duplicidade seja único em todo o dataset
    registros      = []
    grupo_counter  = 0   # ← GLOBAL, fora do loop

    for (_, _), grp in work.groupby(["_cliente_norm", "_servico_norm"], sort=False):
        if len(grp) < 2:
            continue

        grp     = grp.sort_values("_data_parsed").reset_index(drop=True)
        datas   = grp["_data_parsed"].tolist()
        row_ids = grp["_row_id"].tolist()
        n       = len(grp)

        classificacao = {}   # local ao grupo (cliente, serviço)
        i = 0

        while i < n:
            rid_i = row_ids[i]

            # DUPLICATA não vira âncora de novo grupo
            if rid_i in classificacao and classificacao[rid_i]["tipo"] == "DUPLICATA":
                i += 1
                continue

            # Busca OS posteriores dentro da janela ainda não classificadas
            duplicatas_j = []
            for j in range(i + 1, n):
                delta = (datas[j] - datas[i]).days
                if delta <= janela_dias:
                    rid_j = row_ids[j]
                    if rid_j not in classificacao:
                        duplicatas_j.append(j)
                else:
                    break

            if duplicatas_j:
                grupo_counter += 1   # ← incrementa o contador GLOBAL

                if rid_i not in classificacao:
                    classificacao[rid_i] = {"grupo": grupo_counter, "tipo": "ORIGINAL"}

                for j in duplicatas_j:
                    classificacao[row_ids[j]] = {"grupo": grupo_counter, "tipo": "DUPLICATA"}

            i += 1

        for rid, info in classificacao.items():
            registros.append({
                "_row_id":           rid,
                "grupo_duplicidade": info["grupo"],
                "tipo_registro":     info["tipo"],
            })

    if not registros:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # ── Monta DataFrame de saída ──────────────────────────────────────────────
    df_class  = pd.DataFrame(registros).drop_duplicates("_row_id")
    df_merged = work.merge(df_class, on="_row_id", how="inner")
    df_merged = df_merged.sort_values(
        ["grupo_duplicidade", "_data_parsed", "tipo_registro"]
    ).reset_index(drop=True)

    # ── Detalhamento ─────────────────────────────────────────────────────────
    output_cols = ["grupo_duplicidade", "tipo_registro"]
    if col_os:
        output_cols.append(col_os)
    output_cols += [col_cliente, col_servico, col_data]
    if cols_extras:
        output_cols += [c for c in cols_extras if c not in output_cols]

    df_det = df_merged[[c for c in output_cols if c in df_merged.columns]].copy()
    df_det[col_data] = df_merged["_data_parsed"].dt.strftime("%d/%m/%Y")

    # ── Resumo por grupo ──────────────────────────────────────────────────────
    resumo_grupos = []
    for gid, grp in df_merged.groupby("grupo_duplicidade"):
        orig = grp[grp["tipo_registro"] == "ORIGINAL"]
        dups = grp[grp["tipo_registro"] == "DUPLICATA"]
        if orig.empty:
            continue

        o         = orig.iloc[0]
        datas_dup = dups["_data_parsed"]

        row = {
            "grupo":                   int(gid),
            "cliente":                 o[col_cliente],
            "tipo_servico":            o[col_servico],
            "data_os_original":        o["_data_parsed"].strftime("%d/%m/%Y"),
            "qtd_duplicatas":          len(dups),
            "data_primeira_duplicata": datas_dup.min().strftime("%d/%m/%Y") if not dups.empty else "—",
            "data_ultima_duplicata":   datas_dup.max().strftime("%d/%m/%Y") if not dups.empty else "—",
            "intervalo_dias":          (datas_dup.max() - datas_dup.min()).days if not dups.empty else 0,
        }
        if col_os:
            row["os_original"]   = str(o[col_os])
            row["os_duplicadas"] = ", ".join(dups[col_os].astype(str).tolist()) if not dups.empty else "—"
        resumo_grupos.append(row)

    df_resumo_grupos = pd.DataFrame(resumo_grupos)

    # ── Resumo por tipo de serviço ────────────────────────────────────────────
    cont_cli_serv = (
        work.groupby(["_servico_norm", "_cliente_norm"])
        .size()
        .reset_index(name="total_os_cliente")
    )

    resumo_servico = []
    dups_only = df_merged[df_merged["tipo_registro"] == "DUPLICATA"]

    for servico_norm, grp_serv in dups_only.groupby("_servico_norm"):
        servico_val  = grp_serv[col_servico].iloc[0]
        total_dup    = grp_serv.shape[0]
        clientes_dup = grp_serv["_cliente_norm"].nunique()
        total_os     = work[work["_servico_norm"] == servico_norm].shape[0]

        dist = cont_cli_serv[
            cont_cli_serv["_servico_norm"] == servico_norm
        ]["total_os_cliente"]

        resumo_servico.append({
            "tipo_servico":                 servico_val,
            "total_os_no_periodo":          int(total_os),
            "total_duplicatas":             int(total_dup),
            "clientes_com_duplicata":       int(clientes_dup),
            "media_duplicatas_por_cliente": round(total_dup / clientes_dup, 2) if clientes_dup else 0,
            "clientes_1_pedido":            int((dist == 1).sum()),
            "clientes_2_pedidos":           int((dist == 2).sum()),
            "clientes_3_pedidos":           int((dist == 3).sum()),
            "clientes_4_a_6_pedidos":       int(((dist >= 4) & (dist <= 6)).sum()),
            "clientes_7_a_10_pedidos":      int(((dist >= 7) & (dist <= 10)).sum()),
            "clientes_mais_10_pedidos":     int((dist > 10).sum()),
        })

    df_resumo_servico = pd.DataFrame(resumo_servico).sort_values(
        "total_duplicatas", ascending=False
    ).reset_index(drop=True)

    return df_det, df_resumo_grupos, df_resumo_servico


# ── Exportação Excel ──────────────────────────────────────────────────────────
def to_excel_bytes(
    df_det: pd.DataFrame,
    df_grupos: pd.DataFrame,
    df_servico: pd.DataFrame,
    janela_dias: int,
) -> bytes:
    wb = Workbook()

    header_fill   = PatternFill("solid", fgColor="1F4E79")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin          = Side(style="thin", color="BFBFBF")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_original = PatternFill("solid", fgColor="C6EFCE")
    palette       = [
        "FFF2CC", "FDEBD0", "D5F5E3", "D6EAF8", "F9EBEA",
        "EAF2FF", "FDF2F8", "E8F8F5", "FDFEFE", "F4ECF7",
    ]

    def write_sheet(ws, df, title, grupo_col=None, destacar_tipo=False):
        ws.title = title
        if df.empty:
            ws.cell(1, 1, "Nenhum dado para exibir.")
            return

        for c_idx, col in enumerate(df.columns, 1):
            cell           = ws.cell(row=1, column=c_idx, value=col)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(c_idx)].width = max(16, len(str(col)) + 6)
        ws.row_dimensions[1].height = 28

        has_tipo = destacar_tipo and "tipo_registro" in df.columns

        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            grupo_val = None
            if grupo_col and grupo_col in df.columns:
                grupo_val = getattr(row, grupo_col.replace(" ", "_"), None)

            tipo_val = getattr(row, "tipo_registro", None) if has_tipo else None

            if tipo_val == "ORIGINAL":
                row_fill = fill_original
            elif grupo_val is not None:
                color    = palette[(int(grupo_val) - 1) % len(palette)]
                row_fill = PatternFill("solid", fgColor=color)
            else:
                row_fill = None

            for c_idx, val in enumerate(row, 1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")
                if row_fill:
                    cell.fill = row_fill

        ws.freeze_panes = "A2"

    ws1 = wb.active
    write_sheet(ws1, df_det, "Detalhamento", grupo_col="grupo_duplicidade", destacar_tipo=True)

    ws2 = wb.create_sheet("Resumo por Grupo")
    write_sheet(ws2, df_grupos, "Resumo por Grupo", grupo_col="grupo")

    ws3 = wb.create_sheet("Resumo por Serviço")
    write_sheet(ws3, df_servico, "Resumo por Serviço")

    ws4 = wb.create_sheet("Configurações")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 35
    total_dup = (
        int(df_det[df_det["tipo_registro"] == "DUPLICATA"].shape[0])
        if not df_det.empty and "tipo_registro" in df_det.columns
        else 0
    )
    meta = [
        ("Gerado em",                datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Janela de análise (dias)", janela_dias),
        ("Total de grupos",          df_grupos["grupo"].nunique() if not df_grupos.empty else 0),
        ("Total de OS duplicadas",   total_dup),
        ("Serviços com duplicidade", df_servico["tipo_servico"].nunique() if not df_servico.empty else 0),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws4.cell(r, 1, k).font = Font(bold=True)
        ws4.cell(r, 2, str(v))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Paginação ─────────────────────────────────────────────────────────────────
def paginar(df: pd.DataFrame, key: str, page_size: int = 500):
    if len(df) <= page_size:
        st.dataframe(df, use_container_width=True, height=450)
        return
    total_pages = (len(df) - 1) // page_size + 1
    page = st.number_input(
        "Página", min_value=1, max_value=total_pages, value=1, step=1, key=key
    )
    start = (page - 1) * page_size
    st.dataframe(df.iloc[start: start + page_size], use_container_width=True, height=450)
    st.caption(
        f"Exibindo {start+1:,}–{min(start+page_size, len(df)):,} de {len(df):,} registros"
    )


# ── Ingestão de novo lote ─────────────────────────────────────────────────────
def ingerir_lote():
    st.sidebar.divider()
    st.sidebar.subheader("📥 Inserir novo lote")

    arq_lote = st.sidebar.file_uploader(
        "Arquivo do lote (.xlsx/.csv)",
        type=["xlsx", "xls", "csv"],
        key="upload_lote",
    )

    if not arq_lote:
        return

    buf = arq_lote.read()

    try:
        df_preview = (
            pd.read_csv(io.BytesIO(buf), dtype=str, nrows=3)
            if arq_lote.name.endswith(".csv")
            else pd.read_excel(io.BytesIO(buf), dtype=str, nrows=3)
        )
    except Exception as e:
        st.sidebar.error(f"Erro ao ler arquivo: {e}")
        return

    df_preview.columns = df_preview.columns.str.strip()
    cols_lote = list(df_preview.columns)

    lote_nome    = st.sidebar.text_input("Nome do lote", value=datetime.now().strftime("%Y-%m"))
    lote_cliente = st.sidebar.selectbox("Coluna matricula",       cols_lote, key="lc")
    lote_servico = st.sidebar.selectbox("Coluna tipo de serviço",       cols_lote, key="ls")
    lote_data    = st.sidebar.selectbox("Coluna data",          cols_lote, key="ld")
    lote_os_raw  = st.sidebar.selectbox("Coluna OS (opcional)", ["— nenhuma —"] + cols_lote, key="lo")
    lote_os      = None if lote_os_raw == "— nenhuma —" else lote_os_raw

    if not st.sidebar.button("⚙️ Processar e adicionar à base", use_container_width=True):
        return

    try:
        df_lote = (
            pd.read_csv(io.BytesIO(buf), dtype=str)
            if arq_lote.name.endswith(".csv")
            else pd.read_excel(io.BytesIO(buf), dtype=str)
        )
        df_lote.columns = df_lote.columns.str.strip()
    except Exception as e:
        st.sidebar.error(f"Erro ao processar arquivo: {e}")
        return

    obrigatorias = [lote_cliente, lote_servico, lote_data]
    faltando = [c for c in obrigatorias if c not in df_lote.columns]
    if faltando:
        st.sidebar.error(f"Colunas não encontradas: {faltando}")
        return

    lotes_existentes = listar_lotes()
    if lote_nome in lotes_existentes:
        st.sidebar.error(f"Lote '{lote_nome}' já existe. Escolha outro nome.")
        return

    df_lote["_data_parsed"]  = parse_dates_robust(df_lote[lote_data].astype(str).str.strip())
    df_lote["_cliente_norm"] = df_lote[lote_cliente].astype(str).str.strip().str.upper()
    df_lote["_servico_norm"] = df_lote[lote_servico].astype(str).str.strip().str.upper()
    df_lote["_lote"]         = lote_nome
    df_lote["_ingestao_ts"]  = datetime.now().isoformat()

    invalidas = df_lote["_data_parsed"].isna().sum()

    if PARQUET_MESTRE.exists():
        df_antigo = pd.read_parquet(PARQUET_MESTRE)
        df_total  = pd.concat([df_antigo, df_lote], ignore_index=True)
    else:
        df_total = df_lote

    df_total.to_parquet(PARQUET_MESTRE, index=False, compression="snappy")

    st.sidebar.success(
        f"✅ Lote **{lote_nome}** adicionado!\n\n"
        f"- {len(df_lote):,} registros no lote\n"
        f"- {len(df_total):,} registros na base total\n"
        f"- {invalidas:,} datas inválidas ignoradas"
    )
    st.cache_data.clear()
    st.rerun()


# ── Interface principal ───────────────────────────────────────────────────────
st.title("🔍 Detector de Duplicidades de OS")
st.caption(
    "Base histórica via DuckDB + Parquet | "
    "A OS original não entra na contagem de duplicatas."
)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("📦 Base Histórica")

    if not PARQUET_MESTRE.exists():
        st.warning("Nenhuma base mestre encontrada.\nIngira o primeiro lote abaixo.")
    else:
        stats = stats_base()
        st.metric("Total de OS",       f"{int(stats.get('total_registros', 0)):,}")
        st.metric("Lotes inseridos",    int(stats.get("total_lotes", 0)))
        st.metric("Clientes únicos",    f"{int(stats.get('clientes_unicos', 0)):,}")
        st.metric("Serviços distintos", int(stats.get("servicos_unicos", 0)))

        data_min = stats.get("data_min", "")
        data_max = stats.get("data_max", "")
        if data_min and data_max:
            try:
                d_min = pd.to_datetime(data_min).strftime("%d/%m/%Y")
                d_max = pd.to_datetime(data_max).strftime("%d/%m/%Y")
                st.caption(f"Período na base: {d_min} → {d_max}")
            except Exception:
                pass

        if st.button("🔄 Recarregar estatísticas", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

ingerir_lote()

st.divider()

# ── Verifica base ─────────────────────────────────────────────────────────────
if not PARQUET_MESTRE.exists():
    st.info("Insira o primeiro lote pela barra lateral para começar.")
    st.stop()

# ── Filtros de consulta ───────────────────────────────────────────────────────
st.subheader("🗓️ Filtros de consulta")

lotes_disponiveis = listar_lotes()

fc1, fc2 = st.columns(2)
usar_periodo = fc1.checkbox("Filtrar por período de datas", value=True)
usar_lotes   = fc2.checkbox("Filtrar por lote(s) específico(s)")

data_ini_q = data_fim_q = None
lotes_sel  = None

if usar_periodo:
    p1, p2 = st.columns(2)
    data_ini_q = str(p1.date_input(
        "Data inicial", value=datetime.today() - timedelta(days=90)
    ))
    data_fim_q = str(p2.date_input(
        "Data final", value=datetime.today()
    ))

if usar_lotes:
    lotes_sel = st.multiselect(
        "Lotes",
        lotes_disponiveis,
        default=lotes_disponiveis[-1:] if lotes_disponiveis else [],
    )

if st.button("📥 Carregar dados da base", use_container_width=True):
    df_base = carregar_base(
        data_ini=data_ini_q,
        data_fim=data_fim_q,
        lotes=lotes_sel if usar_lotes else None,
    )
    st.session_state["df_base"]   = df_base
    st.session_state["cols_base"] = [c for c in df_base.columns if not c.startswith("_")]
    st.success(f"✅ {len(df_base):,} registros carregados.")

# ── Configurações de análise ──────────────────────────────────────────────────
if "df_base" in st.session_state and not st.session_state["df_base"].empty:
    df_base       = st.session_state["df_base"]
    cols_visiveis = st.session_state["cols_base"]

    with st.expander("👁️ Prévia dos dados carregados (10 primeiras linhas)", expanded=False):
        st.dataframe(df_base[cols_visiveis].head(10), use_container_width=True)

    st.divider()
    st.subheader("⚙️ Configurações da análise")

    def suggest(keywords, cols):
        for kw in keywords:
            for col in cols:
                if kw.lower() in col.lower():
                    return col
        return cols[0] if cols else None

    ca1, ca2, ca3 = st.columns(3)

    col_cliente = ca1.selectbox(
        "👤 Coluna de matrícula",
        cols_visiveis,
        index=cols_visiveis.index(
            suggest(["cliente", "matricula", "cpf", "cod"], cols_visiveis) or cols_visiveis[0]
        ),
    )
    col_servico = ca2.selectbox(
        "🔧 Coluna de tipo de serviço",
        cols_visiveis,
        index=cols_visiveis.index(
            suggest(["servico", "serviço", "tipo", "categoria"], cols_visiveis) or cols_visiveis[0]
        ),
    )
    col_data = ca3.selectbox(
        "📅 Coluna de data",
        cols_visiveis,
        index=cols_visiveis.index(
            suggest(["data", "dt_", "abertura", "criacao"], cols_visiveis) or cols_visiveis[0]
        ),
    )

    ca4, ca5 = st.columns(2)
    none_opt = "— nenhuma —"

    col_os_raw = ca4.selectbox(
        "🔢 Coluna de número da OS (opcional)",
        [none_opt] + cols_visiveis,
    )
    col_os = None if col_os_raw == none_opt else col_os_raw

    janela_dias = ca5.number_input(
        "📆 Janela de duplicidade (dias)",
        min_value=1, max_value=3650, value=30, step=1,
    )

    extras_disp = [
        c for c in cols_visiveis
        if c not in [col_cliente, col_servico, col_data, col_os]
    ]
    cols_extras = st.multiselect("➕ Colunas extras no relatório (opcional)", extras_disp)

    st.divider()

    if st.button("🚀 Analisar duplicidades", type="primary", use_container_width=True):
        df_det, df_grupos, df_servico = detect_duplicates(
            df=df_base,
            col_cliente=col_cliente,
            col_servico=col_servico,
            col_data=col_data,
            janela_dias=int(janela_dias),
            col_os=col_os,
            cols_extras=tuple(cols_extras) if cols_extras else None,
        )
        st.session_state["df_det"]     = df_det
        st.session_state["df_grupos"]  = df_grupos
        st.session_state["df_servico"] = df_servico
        st.session_state["janela"]     = int(janela_dias)

    # ── Resultados ────────────────────────────────────────────────────────────
    if "df_det" in st.session_state:
        df_det     = st.session_state["df_det"]
        df_grupos  = st.session_state["df_grupos"]
        df_servico = st.session_state["df_servico"]
        janela     = st.session_state["janela"]

        st.divider()

        if df_det.empty:
            st.success("✅ Nenhuma duplicidade encontrada com os parâmetros informados.")
        else:
            total_dup = (
                int(df_det[df_det["tipo_registro"] == "DUPLICATA"].shape[0])
                if "tipo_registro" in df_det.columns else 0
            )

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Grupos duplicados",        f"{df_grupos['grupo'].nunique():,}")
            k2.metric("Total de OS duplicadas",   f"{total_dup:,}")
            k3.metric("Janela utilizada",          f"{janela} dias")
            k4.metric("Serviços com duplicidade",
                      f"{df_servico['tipo_servico'].nunique():,}"
                      if not df_servico.empty else "0")

            st.divider()

            tab1, tab2, tab3 = st.tabs([
                "📄 Detalhamento por OS",
                "📊 Resumo por Grupo",
                "🧩 Visão por Tipo de Serviço",
            ])

            with tab1:
                st.caption(
                    "🟢 **Verde** = OS ORIGINAL (legítima, não contada como duplicata) | "
                    "🟡 **Colorido por grupo** = OS DUPLICATA"
                )
                paginar(df_det, key="pag_det")

            with tab2:
                st.caption(
                    "Uma linha por grupo. "
                    "**qtd_duplicatas** não inclui a OS original."
                )
                paginar(df_grupos, key="pag_grp")

            with tab3:
                st.caption(
                    "Agrupamento por tipo de serviço — total de duplicatas, "
                    "clientes afetados e distribuição por faixas de pedidos por cliente."
                )
                if not df_servico.empty:
                    st.dataframe(df_servico, use_container_width=True, height=400)

                    st.divider()
                    st.markdown("#### 📈 Distribuição de pedidos por cliente (top 10 serviços)")

                    faixas = [
                        "clientes_1_pedido",
                        "clientes_2_pedidos",
                        "clientes_3_pedidos",
                        "clientes_4_a_6_pedidos",
                        "clientes_7_a_10_pedidos",
                        "clientes_mais_10_pedidos",
                    ]
                    labels  = ["1", "2", "3", "4–6", "7–10", ">10"]
                    top10   = df_servico.head(10)
                    df_chart = (
                        top10.set_index("tipo_servico")[faixas]
                        .rename(columns=dict(zip(faixas, labels)))
                        .T
                    )
                    st.bar_chart(df_chart)

            st.divider()

            xlsx_bytes = to_excel_bytes(df_det, df_grupos, df_servico, janela)
            st.download_button(
                label="⬇️ Baixar relatório XLSX",
                data=xlsx_bytes,
                file_name=f"duplicidades_os_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
